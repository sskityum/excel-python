
from openpyxl import load_workbook
from openpyxl import Workbook
from xlrd import open_workbook
from xlwt import Workbook

wb_form = load_workbook(filename = 'test.xlsx')
wb_val = load_workbook(filename = 'test.xlsx', data_only=True)

sheet_form1 = wb_form['list1']
sheet_val1 = wb_val['list1']

sheet_form2 = wb_form['list2']
sheet_val2 = wb_val['list2']

D6_form1 = sheet_form1['D6'].value
D6_val1 = sheet_val1['D6'].value

E6_form2 = sheet_form2['E6'].value
E6_val2 = sheet_val2['E6'].value

# print('формула ячейки D6: ' + str(D6_form1) + ',  значение ячейки D6: ' + str(D6_val1))
# print('формула ячейки D6: ' + str(E6_form2) + ',  значение ячейки D6: ' + str(E6_val2))

wb_val_write = Workbook()
wb_val_write.create_sheet(title = 'list3')
wb_val_write.save('test.xlsx')